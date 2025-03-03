import pandas as pd
import numpy as np
from langchain.prompts import ChatPromptTemplate, MessagesPlaceholder
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from dotenv import load_dotenv
from tqdm import tqdm
import warnings
from concurrent.futures import ThreadPoolExecutor, as_completed
from langchain.prompts import ChatPromptTemplate, MessagesPlaceholder
from langchain_anthropic import ChatAnthropic
from langchain_openai import ChatOpenAI
from langchain.callbacks import AsyncIteratorCallbackHandler
from IPython.display import display, HTML
from transformers import GPT2Tokenizer
from pydantic import BaseModel, Field
from typing import List
from openpyxl import Workbook, load_workbook
from tqdm import tqdm
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from concurrent.futures import ThreadPoolExecutor, as_completed
import re
import matplotlib.pyplot as plt
import streamlit as st
 
ANTHROPIC_API_KEY= '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'
 
warnings.filterwarnings("ignore")
 

def format_input_clients(dataframe):
    # Combiner toutes les lignes en une seule chaîne de caractères
    combined_input = " ".join(
        f"Hour: {row['Hour']}, Client_Entries: {row['Client_Entries']}"
        #for _, row in dataframe.sample(500).iterrows()
        for _, row in dataframe.iterrows()
    )
    return combined_input


def format_input_employees(dataframe):
    # Combiner toutes les lignes en une seule chaîne de caractères
    combined_input = " ".join(
        f"Hour: {row['Hour']}, Employee_Count: {row['Employee_Count']}"
        for _, row in dataframe.iterrows()
    )
    return combined_input


def format_input_ca(dataframe):
    # Combiner toutes les lignes en une seule chaîne de caractères
    combined_input = " ".join(
        f"Hour: {row['Hour']}, CA: {row['CA']}"
        for _, row in dataframe.iterrows()
    )
    return combined_input
 

def format_input_clients_ex(dataframe):
    # Combiner toutes les lignes en une seule chaîne de caractères
    combined_input = " ".join(
        f"Date: {row['Date']}, Hour: {row['Hour']}, Customer_Count: {row['Customer_Count']}"
        #for _, row in dataframe.sample(500).iterrows()
        for _, row in dataframe.iterrows()
    )
    return combined_input
 
def format_input_employees_ex(dataframe):
    # Combiner toutes les lignes en une seule chaîne de caractères
    combined_input = " ".join(
        f"Date: {row['Date']}, Hour: {row['Hour']}, Employee_Count: {row['Employee_Count']}"
        for _, row in dataframe.iterrows()
    )
    return combined_input
 
def format_input_ca_ex(dataframe):
    # Combiner toutes les lignes en une seule chaîne de caractères
    combined_input = " ".join(
        f"Hour: {row['Hour']}, CA: {row['CA']}"
        #for _, row in dataframe.sample(500).iterrows()
        for _, row in dataframe.iterrows()
    )
    return combined_input

# Exemples d'inputs et outputs pour entraîner le modèle
extraction_examples = [
        {
            "employees": format_input_employees_ex(pd.read_csv('multiple_days_hourly_employee_counts.csv')),
            "clients": format_input_clients_ex(pd.read_csv('multiple_days_hourly_customer_counts.csv')),
            "ca": format_input_ca_ex(pd.read_csv('chiffre_affaire_par_heure.csv')),
           
            "output": """
                Pour chaque heure de la journée: le nombre d'employés moyen par heure recommandé
 
                8:00 - Recommandé: 8
                9:00 - Recommandé: 6
                10:00 - Recommandé: 7
                11:00 - Recommandé: 7
                12:00 - Recommandé: 8
                13:00 - Recommandé: 7
                14:00 - Recommandé: 4
                15:00 - Recommandé: 5
                16:00 - Recommandé: 7
                17:00 - Recommandé: 3
                18:00 - Recommandé: 5
                19:00 - Recommandé: 6
                20:00 - Recommandé: 7
            """
        },
]
 
# Conversion des exemples en messages
def generate_extraction_messages(examples):
    messages = []
    for example in examples:
        clients_example = example["clients"]
        ca_example = example["ca"]
        employees_example = example["employees"]
        output_example = example["output"]
        messages.append(
            ("human", f"En utilisant le fichier ci dessous qui représente le nombre d'employés par heure en magasin: \n\n<employees>{employees_example}</employees>\n\n, le fichier ci-dessous qui représente le chiffre d'affaire par heure pour le magasin \n\n<ca>{ca_example}</ca>\n\n ,et le fichier ci-dessous qui représente le nombre de clients par heure en magasin: \n\n<client>{clients_example}</client>\n\n DDonne moi pour chaque heure de la journée: le nombre d'employés recommandé.")
        )
        messages.append(
            ("assistant", output_example)
        )
    return messages
 
def llm_anthropic(model: str, temperature: int = 0):
    #callback = AsyncIteratorCallbackHandler()
    llm_init = ChatAnthropic(
        api_key=ANTHROPIC_API_KEY,
        temperature=temperature,
        model=model,
        streaming=True,
        #callbacks=[callback],
    )
    return llm_init
 
def llm_prompt(employees, clients, turnover):
    extraction_messages = generate_extraction_messages(extraction_examples)
    
    model = llm_anthropic('claude-3-5-sonnet-20240620', temperature=0.7)

    # Define the prompt template with examples
    enrichment_prompt_with_examples = ChatPromptTemplate.from_messages(
        [
        (
                "system",
                """
                You are an expert in retail with 20+ years of experience.
                You are working in the HR department and you are responsible for managing workforce management.
                Here workforce management is the planning of the employees' presence in the stores throughout the days.
                Your goal is to adapt the planning to the frequentation of clients in the stores.
                Ensures that the sum of the list of recommended employees equal the sum of the list of current employees.
                """,
            ),
            # MessagesPlaceholder("examples"),
            (
                "human",
                """
                Using the file below:
    
                <context>
                the number of employees:
                    {employees}
                </context>
    
                <context>
                the turnover:
                    {turnover}
                </context>
    
                <context>
                the number of clients:
                    {clients}
                </context>
    
                <context>
                    One of France largest clothes retailer
                </context>
    
                Give me the recommended number of employees for each hour of the day. It must have the same format as {employees}.
                Ensures that the total number of employees on the day remains the same (sum of {employees} = sum of the output list).
                Note that there might be cases where you may not find enough information to describe the relation comprehensively. In such cases, your output should be "Pas assez d'information".
                """
            ),
        ]
    )
    
    class AllocationOutput(BaseModel):
        """Output of the allocation process."""
        employes_recommandes: List[int] = Field(
            ...,
        description="Liste du nombre recommandé d'employés pour chaque heure. La somme de cet output doit être égale à la somme de {employees}"
        )
        desc: str = Field(
            ...,
            description=" Explique ton raisonnement étape par étape et rappelle les contraintes que tu as pris en compte")
        desc_1: str = Field(
            ...,    
        description=" Donne une interpretation de l'ensemble des listes, c'est à dire, explique quel est l'état actuel du nombre d'employés dans le magasin, les problèmes au niveau de l'état actuel et pourquoi tu as proposé un planning différent.")
    
    chain = enrichment_prompt_with_examples | model.with_structured_output(
        schema=AllocationOutput,
        method="function_calling",
        include_raw=False,
    )
    
    result = chain.invoke(
                {
                    "employees": employees
                    ,"clients": clients
                    ,"turnover": turnover
                    ,"examples": extraction_messages
                }
            )

    employes_recommandes = result.employes_recommandes
    desc = result.desc
    desc_1 = result.desc_1

    return employes_recommandes, desc, desc_1

 
def reco(CA_df, employee_df, client_df):
    
    # Extraire la liste des heures
    heures_list = CA_df['Hour'].tolist()
    turnover_list  = CA_df['CA'].tolist()
    clients_list  = client_df['Client_Entries'].tolist()
    employes_actuels_list  = employee_df['Employee_Count'].tolist()
    
    #erreur ici Ca doit pas etre la liste client mais ca doit être le fichier 'client_velizy' ?

    clients = format_input_clients(client_df)
    employees = format_input_employees(employee_df)
    turnover = format_input_ca(CA_df)

    with st.spinner('Generating recommendations ...'):
        employes_recommandes, raisonnement, interpretation = llm_prompt(employees, clients, turnover) #
    print(employes_recommandes, employes_actuels_list)

    heures_reversed = heures_list [::-1]
    employes_actuels_reversed = employes_actuels_list [::-1]
    clients_reversed = clients_list [::-1]
    chiffre_d_affaire_reversed = turnover_list [::-1]
    employes_recommandes_reversed = employes_recommandes[::-1]
    # Configuration du graphique
    fig, ax1 = plt.subplots(figsize=(8, 4))
    
    # Paramètres pour les barres
    bar_width = 0.4  # Largeur des barres
    index = np.arange(len(heures_reversed))  # Positions des heures

    # Barres pour le nombre d'employés actuels
    ax1.barh(index - bar_width/2, employes_actuels_reversed, bar_width, label='Employés Actuels', color='lightgreen')
    
    # Barres pour le nombre recommandé d'employés
    ax1.barh(index + bar_width/2, employes_recommandes_reversed, bar_width, label='Employés Recommandés', color='#006400')
    
    # Ajouter des étiquettes et un titre
    ax1.set_ylabel('Heure', fontsize=8)
    ax1.set_xlabel('Nombre d\'Employés', fontsize=8)
    ax1.set_title('Comparaison du Nombre d\'Employés Actuels et Recommandés par Heure', fontsize=10)
    ax1.set_yticks(index)
    ax1.set_yticklabels(heures_reversed)
    ax1.tick_params(axis='x', labelsize=8)  # Réduire la taille des chiffres sur l'axe x
    ax1.tick_params(axis='y', labelsize=8)  # Réduire la taille des chiffres sur l'axe x
    ax1.legend(loc='lower left', fontsize=6)
    
    # Création d'un second axe pour le CA
    ax2 = ax1.twiny()  # Ajout d'un axe x supplémentaire partagé avec ax1
    ax2.plot(chiffre_d_affaire_reversed, index, color='blue', marker='o', label='Chiffre d\'Affaires (CA) en euros', linewidth=1)
    ax2.set_xlabel('CA (Chiffre d\'Affaires) en euros', fontsize=8)
    ax2.tick_params(axis='x', labelsize=8)
    ax2.legend(loc='lower right', fontsize=6)  # Déplacement de la légende pour le CA
    
    # Création d'un troisième axe pour les clients
    ax3 = ax1.twiny()  # Partage le même axe y
    ax3.spines["top"].set_position(("outward", 30))  # Déplacer légèrement l'axe vers le haut
    ax3.plot(clients_reversed, index, color='red', marker='x', label='Nombre de Clients', linewidth=1)
    ax3.set_xlabel('Nombre de Clients', fontsize=8)
    ax3.tick_params(axis='x', labelsize=8)
    ax3.legend(loc='upper right', fontsize=6) #bbox_to_anchor=(1, 0)

    # Combine all legends in one place
    """lines_labels = [ax1.get_legend_handles_labels(), ax2.get_legend_handles_labels(), ax3.get_legend_handles_labels()]
    lines, labels = [sum(lol, []) for lol in zip(*lines_labels)]
    fig.legend(lines, labels, loc='upper center', bbox_to_anchor=(0.5, 1.1), ncol=3, fontsize=6)"""
    
    # Affichage du graphique
    ax1.grid(axis='x')
    plt.tight_layout()
    """
    print("Heures:", heures_list )
    print("Employés Actuels:", employes_actuels_list )
    print("clients:", clients_list )
    print("turnover:", turnover_list )
    print("Employés Recommandés:", employes_recommandes)"""
    print(sum(employes_actuels_list), sum(employes_recommandes))
    return fig, raisonnement, interpretation
